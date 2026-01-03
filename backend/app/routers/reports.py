"""Reports router - راوتر التقارير."""
from datetime import datetime
from typing import Optional, List, Dict, Any
from uuid import UUID
import io
import json

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.database import get_db
from app.models.assessment import Assessment, AssessmentResponse
from app.models.ndi import NDIDomain, NDIQuestion, NDIMaturityLevel
from app.models.evidence import Evidence
from app.services.score_service import ScoreService

router = APIRouter()


async def get_assessment_with_details(
    assessment_id: UUID,
    db: AsyncSession,
) -> Dict[str, Any]:
    """Get assessment with all details for report."""
    result = await db.execute(
        select(Assessment).where(Assessment.id == assessment_id)
    )
    assessment = result.scalar_one_or_none()

    if not assessment:
        raise HTTPException(status_code=404, detail="Assessment not found")

    # Get scores
    service = ScoreService(db)
    maturity = await service.calculate_maturity_score(assessment_id)
    compliance = await service.calculate_compliance_score(assessment_id)

    # Get responses with questions
    responses_result = await db.execute(
        select(AssessmentResponse)
        .options(
            selectinload(AssessmentResponse.question)
            .selectinload(NDIQuestion.domain),
            selectinload(AssessmentResponse.question)
            .selectinload(NDIQuestion.maturity_levels),
            selectinload(AssessmentResponse.evidence),
        )
        .where(AssessmentResponse.assessment_id == assessment_id)
    )
    responses = responses_result.scalars().all()

    # Build response details
    response_details = []
    for r in responses:
        if r.question:
            # Get selected level details
            selected_level_details = None
            if r.selected_level is not None:
                for ml in r.question.maturity_levels:
                    if ml.level == r.selected_level:
                        selected_level_details = {
                            "level": ml.level,
                            "name_en": ml.name_en,
                            "name_ar": ml.name_ar,
                            "description_en": ml.description_en,
                            "description_ar": ml.description_ar,
                        }
                        break

            response_details.append({
                "question_code": r.question.code,
                "domain_code": r.question.domain.code if r.question.domain else None,
                "domain_name_en": r.question.domain.name_en if r.question.domain else None,
                "domain_name_ar": r.question.domain.name_ar if r.question.domain else None,
                "question_en": r.question.question_en,
                "question_ar": r.question.question_ar,
                "selected_level": r.selected_level,
                "selected_level_details": selected_level_details,
                "justification": r.justification,
                "notes": r.notes,
                "evidence_count": len(r.evidence),
                "evidence": [
                    {
                        "file_name": e.file_name,
                        "file_type": e.file_type,
                        "analysis_status": e.analysis_status,
                    }
                    for e in r.evidence
                ],
            })

    return {
        "assessment": {
            "id": str(assessment.id),
            "name": assessment.name,
            "description": assessment.description,
            "assessment_type": assessment.assessment_type,
            "status": assessment.status,
            "target_level": assessment.target_level,
            "created_at": assessment.created_at.isoformat() if assessment.created_at else None,
            "completed_at": assessment.completed_at.isoformat() if assessment.completed_at else None,
        },
        "maturity": maturity.model_dump(),
        "compliance": compliance.model_dump(),
        "responses": response_details,
    }


@router.get("/assessments/{assessment_id}/report")
async def generate_report(
    assessment_id: UUID,
    format: str = Query("json", pattern="^(json|pdf|excel)$"),
    language: str = Query("ar", pattern="^(ar|en)$"),
    db: AsyncSession = Depends(get_db),
):
    """Generate assessment report in specified format."""
    report_data = await get_assessment_with_details(assessment_id, db)

    # Add metadata
    report_data["generated_at"] = datetime.utcnow().isoformat()
    report_data["language"] = language

    if format == "json":
        return report_data

    elif format == "excel":
        # Generate Excel report
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            wb = openpyxl.Workbook()

            # Summary sheet
            ws = wb.active
            ws.title = "Summary" if language == "en" else "الملخص"

            # Header styling
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

            # Write summary
            ws["A1"] = "Assessment Report" if language == "en" else "تقرير التقييم"
            ws["A1"].font = Font(bold=True, size=16)

            ws["A3"] = "Overall Maturity Score:" if language == "en" else "درجة النضج الإجمالية:"
            ws["B3"] = report_data["maturity"]["overall_score"]

            ws["A4"] = "Maturity Level:" if language == "en" else "مستوى النضج:"
            ws["B4"] = report_data["maturity"]["overall_level_name_en" if language == "en" else "overall_level_name_ar"]

            ws["A5"] = "Compliance Percentage:" if language == "en" else "نسبة الامتثال:"
            ws["B5"] = f"{report_data['compliance']['compliance_percentage']}%"

            ws["A7"] = "Generated At:" if language == "en" else "تاريخ الإنشاء:"
            ws["B7"] = report_data["generated_at"]

            # Domain scores sheet
            ws2 = wb.create_sheet("Domain Scores" if language == "en" else "درجات المجالات")

            headers = ["Domain Code", "Domain Name", "Score", "Level", "Answered", "Total"]
            if language == "ar":
                headers = ["كود المجال", "اسم المجال", "الدرجة", "المستوى", "المجاب", "الإجمالي"]

            for col, header in enumerate(headers, 1):
                cell = ws2.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill

            for row, ds in enumerate(report_data["maturity"]["domain_scores"], 2):
                ws2.cell(row=row, column=1, value=ds["domain_code"])
                ws2.cell(row=row, column=2, value=ds["domain_name_en" if language == "en" else "domain_name_ar"])
                ws2.cell(row=row, column=3, value=ds["score"])
                ws2.cell(row=row, column=4, value=ds["level_name_en" if language == "en" else "level_name_ar"])
                ws2.cell(row=row, column=5, value=ds["answered_count"])
                ws2.cell(row=row, column=6, value=ds["total_questions"])

            # Responses sheet
            ws3 = wb.create_sheet("Responses" if language == "en" else "الإجابات")

            headers = ["Question Code", "Domain", "Question", "Selected Level", "Evidence Count"]
            if language == "ar":
                headers = ["كود السؤال", "المجال", "السؤال", "المستوى المختار", "عدد الشواهد"]

            for col, header in enumerate(headers, 1):
                cell = ws3.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill

            for row, resp in enumerate(report_data["responses"], 2):
                ws3.cell(row=row, column=1, value=resp["question_code"])
                ws3.cell(row=row, column=2, value=resp["domain_code"])
                ws3.cell(row=row, column=3, value=resp["question_en" if language == "en" else "question_ar"])
                ws3.cell(row=row, column=4, value=resp["selected_level"])
                ws3.cell(row=row, column=5, value=resp["evidence_count"])

            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f"attachment; filename=ndi_report_{assessment_id}.xlsx"
                },
            )

        except ImportError:
            raise HTTPException(
                status_code=500,
                detail="Excel generation requires openpyxl. Install it with: pip install openpyxl"
            )

    elif format == "pdf":
        # For PDF, return instructions or implement with reportlab/weasyprint
        raise HTTPException(
            status_code=501,
            detail="PDF generation not yet implemented. Use JSON or Excel format."
        )


@router.get("/assessments/{assessment_id}/export")
async def export_assessment_data(
    assessment_id: UUID,
    db: AsyncSession = Depends(get_db),
):
    """Export assessment data for backup or transfer."""
    report_data = await get_assessment_with_details(assessment_id, db)

    # Convert to JSON bytes
    json_bytes = json.dumps(report_data, ensure_ascii=False, indent=2).encode('utf-8')
    output = io.BytesIO(json_bytes)

    return StreamingResponse(
        output,
        media_type="application/json",
        headers={
            "Content-Disposition": f"attachment; filename=ndi_assessment_{assessment_id}.json"
        },
    )


@router.get("/summary")
async def get_reports_summary(
    db: AsyncSession = Depends(get_db),
):
    """Get summary of all completed assessments."""
    result = await db.execute(
        select(Assessment)
        .where(Assessment.status == "completed")
        .order_by(Assessment.completed_at.desc())
    )
    assessments = result.scalars().all()

    summaries = []
    service = ScoreService(db)

    for assessment in assessments:
        maturity = await service.calculate_maturity_score(assessment.id)
        compliance = await service.calculate_compliance_score(assessment.id)

        summaries.append({
            "id": str(assessment.id),
            "name": assessment.name,
            "completed_at": assessment.completed_at.isoformat() if assessment.completed_at else None,
            "maturity_score": maturity.overall_score,
            "maturity_level": maturity.overall_level,
            "maturity_level_name_en": maturity.overall_level_name_en,
            "maturity_level_name_ar": maturity.overall_level_name_ar,
            "compliance_percentage": compliance.compliance_percentage,
            "is_compliant": compliance.is_compliant,
        })

    return {"completed_assessments": summaries}
